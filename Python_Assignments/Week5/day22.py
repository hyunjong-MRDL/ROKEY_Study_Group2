from pathlib import Path

root = Path.cwd()  # Current Working Directory
for item in root.iterdir():
    if item.suffix == ".py":
        print(item)

new_file = Path("practice.py")
if not new_file.exists():
    new_file.touch()   # 새 파일 만들기

if new_file.exists():
    new_file.unlink()  # 기존 파일 삭제

from openpyxl import Workbook, load_workbook
xl_filename = r"example.xlsx"
wb = Workbook()   # 새로운 Excel 파일 생성
ws = wb.active    # 활성된 시트 불러오기

data = [
    ['과일', '가격'],
    ['사과', 1000],
    ['바나나', 800],
    ['딸기', 2000]
]

for row in data:
    ws.append(row) # openpyxl에서 데이터 삽입

wb.save(xl_filename)

wb2 = load_workbook(xl_filename)  # 기존에 존재하는 엑셀 읽기
ws2 = wb2.active                  # 활성 시트 불러오기 (생략하면 안 됨)
ws2.append(['수박', 1500])
wb2.save(xl_filename)

xl_path = Path(xl_filename)
if xl_path.exists():
    xl_path.unlink()